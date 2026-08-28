from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent

HEADERS = [
    "Alert Name",
    "PPL Query",
    "Description",
    "Affected Services",
    "Severity Level",
    "Action for NOC",
    "Contact Info",
    "Alert Frequency",
]

COLUMN_WIDTHS = [42, 80, 60, 60, 14, 80, 60, 18]

CONTACT_INFO = (
    "Primary: <CONTACT_NAME> — <CONTACT_EMAIL> — <CONTACT_PHONE>\n"
    "Team: <TEAM_NAME>\n"
    "Notification group: <NOTIFICATION_GROUP>"
)

ALERT_FREQUENCY = "Every 10 minutes"
SEVERITY = "Critical"

PPL_TEMPLATE = (
    "source=<OPENSEARCH_LOG_INDEX_PATTERN>\n"
    "| where `resource.attributes.service.name` = '{service_name}'\n"
    "| where time >= date_sub(now(), INTERVAL 15 MINUTE)\n"
    "| where severityText = 'ERROR'\n"
    "| spath input=body path=process output=process\n"
    "| where process LIKE 'is_healthy%'\n"
    "| stats count() as count by process\n"
    "| where count > 0\n"
    "| sort - count"
)

ROWS = [
    {
        "alert_name": "<ALERT_NAME>",
        "service_name": "<OTEL_SERVICE_NAME>",
        "microservice_slug": "<microservice_slug>",
        "description": (
            "Triggered when the /monitor endpoint of <microservice_slug> reports "
            "any external dependency as unhealthy (<comma-separated dependency labels>)."
        ),
        "affected_services": (
            "<one-line description of what stage of the pipeline is broken when this microservice is down>."
        ),
        "url": "<PRODUCTION_URL_BASE>/api/monitor/applicative-health-check",
        "openshift_app": "<OPENSHIFT_APP>",
    },
]


def build_action_for_noc(url, openshift_app):
    return (
        "1. Verify the failure by calling the endpoint:\n"
        f"   curl {url}\n"
        '   Look for any element with "is_ok": false; "service_name" identifies the failed dependency; '
        '"description" gives the error.\n'
        "2. Restart the pod in OpenShift:\n"
        "   Project: <OPENSHIFT_PROJECT>\n"
        f"   App:     {openshift_app}\n"
        "3. Re-run the curl after the pod is Ready (~30 sec).\n"
        "4. If still failing after 2 restart cycles, escalate to:\n"
        "   <CONTACT_NAME> — <CONTACT_EMAIL> — <CONTACT_PHONE>"
    )


def build_workbook(row_data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Alert"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    for column_index, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for column_index, width in enumerate(COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    values = [row_data["alert_name"], PPL_TEMPLATE.format(service_name=row_data["service_name"]), row_data["description"], row_data["affected_services"], SEVERITY, build_action_for_noc(row_data["url"], row_data["openshift_app"]), CONTACT_INFO, ALERT_FREQUENCY]
    for column_index, value in enumerate(values, start=1):
        cell = worksheet.cell(row=2, column=column_index, value=value)
        cell.alignment = body_alignment
    worksheet.row_dimensions[1].height = 32
    worksheet.row_dimensions[2].height = 220
    worksheet.freeze_panes = "A2"
    return workbook


def main():
    for row_data in ROWS:
        output_file = OUTPUT_DIR / f"{row_data['microservice_slug']}_alerts.xlsx"
        workbook = build_workbook(row_data)
        workbook.save(output_file)
        print(f"Wrote {output_file}")


if __name__ == "__main__":
    main()
